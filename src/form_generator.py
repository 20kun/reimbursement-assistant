"""Excel reimbursement form generator — outputs professional .xlsx ready for submission."""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_excel(form_data: dict, compliance_result: dict | None = None) -> bytes:
    """Generate a formatted reimbursement form as .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "报销单"

    # --- Styles ---
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="微软雅黑", size=16, bold=True)
    body_font = Font(name="微软雅黑", size=10)
    money_font = Font(name="微软雅黑", size=10)
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    error_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    pass_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # --- Column widths ---
    col_widths = {"A": 8, "B": 14, "C": 12, "D": 14, "E": 28, "F": 18, "G": 24, "H": 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # --- Title ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="费  用  报  销  单")
    cell.font = title_font
    cell.alignment = center_align
    row += 2

    # --- Header info ---
    info_fields = [
        ("申请人", form_data.get("申请人", ""), "部门", form_data.get("部门", "")),
        ("报销日期", form_data.get("报销日期", date.today().isoformat()), "单据编号", "（自动生成）"),
        ("报销事由", form_data.get("报销事由", ""), "附件数量", f"{form_data.get('附件数量', 0)} 张"),
    ]

    for left_label, left_val, right_label, right_val in info_fields:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        label_cell = ws.cell(row=row, column=1, value=left_label)
        label_cell.font = Font(name="微软雅黑", size=10, bold=True)
        label_cell.border = thin_border
        label_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        val_cell = ws.cell(row=row, column=3, value=str(left_val))
        val_cell.font = body_font
        val_cell.border = thin_border
        val_cell.alignment = left_align

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        label_cell2 = ws.cell(row=row, column=5, value=right_label)
        label_cell2.font = Font(name="微软雅黑", size=10, bold=True)
        label_cell2.border = thin_border
        label_cell2.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        val_cell2 = ws.cell(row=row, column=7, value=str(right_val))
        val_cell2.font = body_font
        val_cell2.border = thin_border
        val_cell2.alignment = left_align

        row += 1

    row += 1

    # --- Table header ---
    has_reimbursable = any("可报销金额" in item for item in form_data.get("明细", []))
    headers = ["序号", "日期", "类别", "实际金额", "可报销金额", "开票单位", "发票号码", "事由/物品明细", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # Update column widths for new columns
    col_widths = {"A": 6, "B": 12, "C": 10, "D": 12, "E": 12, "F": 22, "G": 16, "H": 22, "I": 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # --- Table body ---
    items = form_data.get("明细", [])
    for item in items:
        actual = float(item.get("实际金额", item.get("金额", 0)))
        reimbursable = float(item.get("可报销金额", actual))
        capped = item.get("是否截断", "否")
        notes_parts = []
        if capped == "是":
            notes_parts.append(f"超限截断 ¥{actual - reimbursable:,.2f}")
        extra = item.get("_extra_fields", {})
        if extra:
            notes_parts.extend(f"{v}" for v in extra.values() if v)

        values = [
            item.get("序号", ""),
            item.get("日期", ""),
            item.get("类别", ""),
            actual,
            reimbursable,
            item.get("开票单位", ""),
            item.get("发票号", ""),
            item.get("事由/物品", ""),
            "；".join(notes_parts) if notes_parts else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            if col_idx == 4:  # amount column
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            elif col_idx in (1, 2):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        row += 1

    # --- Total row ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value="合  计")
    cell.font = Font(name="微软雅黑", size=10, bold=True)
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col in range(2, 4):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
        c.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Actual total
    total_actual = ws.cell(row=row, column=4, value=float(form_data.get("合计金额", 0)))
    total_actual.font = Font(name="微软雅黑", size=10, bold=True)
    total_actual.number_format = '#,##0.00'
    total_actual.alignment = right_align
    total_actual.border = thin_border
    total_actual.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Reimbursable total
    total_reim = ws.cell(row=row, column=5, value=float(form_data.get("合计可报销", 0)))
    total_reim.font = Font(name="微软雅黑", size=10, bold=True)
    total_reim.number_format = '#,##0.00'
    total_reim.alignment = right_align
    total_reim.border = thin_border
    total_reim.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for col in range(6, 10):
        c = ws.cell(row=row, column=col)
        c.border = thin_border
        c.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    row += 2

    # --- Compliance summary ---
    if compliance_result:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        status_label = {
            "pass": "✅ 合规通过",
            "violation": "❌ 存在违规",
            "needs_review": "⚠️ 需补充信息",
        }.get(compliance_result.get("overall_status", ""), "")
        cell = ws.cell(row=row, column=1, value=f"AI审核结果：{status_label}  |  {compliance_result.get('summary', '')}")
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        fill_map = {"pass": pass_fill, "violation": error_fill, "needs_review": warn_fill}
        cell.fill = fill_map.get(compliance_result.get("overall_status", ""), warn_fill)
        cell.alignment = left_align
        cell.border = thin_border
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        for check in compliance_result.get("checks", []):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            icon = {"pass": "✅", "fail": "❌", "warning": "⚠️"}.get(check.get("status", ""), "•")
            detail = f"  {icon} {check.get('rule')}: {check.get('detail')}"
            if check.get("suggestion"):
                detail += f"  → {check.get('suggestion')}"
            cell = ws.cell(row=row, column=1, value=detail)
            cell.font = body_font
            cell.alignment = left_align
            cell.border = thin_border
            for col in range(2, 9):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    row += 2

    # --- Signature area ---
    sig_fields = [
        ("申请人签字", "审批人签字", "财务审核", "出纳"),
    ]
    for sig_row in sig_fields:
        for i, (label, *_) in enumerate([(s,) for s in sig_row]):
            col = i * 2 + 1
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = body_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.cell(row=row, column=col + 1).border = thin_border
        row += 1
        for i in range(4):
            col = i * 2 + 1
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            cell = ws.cell(row=row, column=col, value="日期：")
            cell.font = body_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.cell(row=row, column=col + 1).border = thin_border
        row += 1

    # --- Row height ---
    ws.row_dimensions[1].height = 30
    for r in range(2, row):
        if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < 20:
            ws.row_dimensions[r].height = 22

    # --- Print settings ---
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
